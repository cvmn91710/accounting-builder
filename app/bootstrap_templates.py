"""Create minimal placeholder Excel template and mapping if missing (dev only). Spec v1.2 — single master workbook."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from app.config import get_settings


def _minimal_mapping_v12() -> dict:
    """Starter mapping — verify row/column indices against the live doc 2290 template before production."""
    return {
        "templatePath": "templates/2290_Accounting_Template.xlsx",
        "sheets": {
            "workingBalance": {
                "sheet": "Working Balance",
                "matterNameCell": "A1",
                "caseNumberCell": "A2",
                "accountingTypeCell": "B4",
                "fiduciaryNameCell": "B5",
                "periodCell": "B7",
            },
            "statements": {
                "sheet": "Statements",
                "blockStartColumns": ["B", "I", "P"],
                "headerRow": 2,
                "dateStartRow": 5,
            },
            "bankTransactions": {
                "sheet": "Bank Statement Transactions",
                "firstBlockStartRow": 5,
                "blockGapRows": 2,
                "columns": {
                    "date": "A",
                    "description": "B",
                    "account": "C",
                    "check": "D",
                    "copy_chk": "E",
                    "debit": "F",
                    "credit": "G",
                    "additional_info": "H",
                },
            },
            "scheduleB": {
                "sheet": "POH @ Beginning Schedule B",
                "cashStartRow": 9,
                "nonCashStartRow": 16,
            },
            "scheduleE": {
                "sheet": "POH @ End Schedule E",
                "cashStartRow": 10,
                "nonCashStartRow": 17,
            },
            "scheduleA": {
                "sheet": "Schedule A",
                "subcategories": {
                    "Interest": {"startRow": 9},
                    "Pensions, Annuities, and Other Regular Periodic Payments": {"startRow": 21},
                    "Miscellaneous Receipts": {"startRow": 38},
                },
                "columns": {"date": "A", "description": "B", "amount": "C"},
            },
            "scheduleC": {
                "sheet": "Schedule C",
                "subcategories": {
                    "Residential Facility/Caregiver": {"startRow": 10},
                    "Living Expenses": {"startRow": 29},
                    "Medical Expenses": {"startRow": 49},
                    "Legal & Related Professional Expenses": {"startRow": 66},
                    "Insurance": {"startRow": 79},
                    "Miscellaneous": {"startRow": 92},
                },
                "columns": {
                    "date": "A",
                    "account": "B",
                    "description": "C",
                    "check": "D",
                    "amount": "E",
                },
            },
            "scheduleF": {
                "sheet": "Schedule F",
                "dataStartRow": 7,
                "columns": {"date": "A", "description": "B", "carry_value": "C"},
            },
        },
        "adHocSchedules": {
            "D": {
                "label": "Losses on Sales During the Period",
                "addToWorkingBalance": "CREDITS",
            },
            "G": {
                "label": "Distributions to Beneficiaries / Conservatee / Minor",
                "addToWorkingBalance": "CREDITS",
            },
            "I": {
                "label": "Net Income from Trade or Business",
                "addToWorkingBalance": "CHARGES",
            },
            "K": {
                "label": "Change in Assets",
                "addToWorkingBalance": "CREDITS",
            },
            "L": {
                "label": "Net Loss from Trade or Business",
                "addToWorkingBalance": "CREDITS",
            },
            "P": {
                "label": "Professional Fees",
                "addToWorkingBalance": "CREDITS",
            },
            "X": {"label": "Cash Reconciliation", "addToWorkingBalance": None},
        },
    }


def _build_placeholder_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    cover = wb.active
    cover.title = "Working Balance"
    cover["A1"] = "Matter Name"
    cover["A2"] = "Case No."
    cover["B4"] = "Accounting type"
    cover["B5"] = "Fiduciary"
    cover["B7"] = "Period"

    for title in (
        "Statements",
        "Bank Statement Transactions",
        "POH @ Beginning Schedule B",
        "POH @ End Schedule E",
        "Schedule A",
        "Schedule C",
        "Schedule F",
    ):
        wb.create_sheet(title)

    ws = wb["Bank Statement Transactions"]
    ws["A4"] = "Date"
    ws["B4"] = "Description"
    ws["C4"] = "Acct"
    ws["D4"] = "Chk"
    ws["E4"] = "Copy"
    ws["F4"] = "Debit"
    ws["G4"] = "Credit"
    ws["H4"] = "Notes"

    wb.save(path)


def ensure_placeholder_templates() -> None:
    s = get_settings()
    mapping_path = s.template_mapping_path
    if not mapping_path.parent.exists():
        mapping_path.parent.mkdir(parents=True, exist_ok=True)
    if not mapping_path.exists():
        mapping_path.write_text(
            json.dumps(_minimal_mapping_v12(), indent=2), encoding="utf-8"
        )

    tpl = s.template_path
    if not tpl.exists():
        tpl.parent.mkdir(parents=True, exist_ok=True)
        _build_placeholder_workbook(tpl)
