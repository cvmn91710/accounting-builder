"""Populate firm Excel template and append hidden audit trail sheet — legacy + spec v1.2."""

from __future__ import annotations

import warnings
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from app.config import get_settings, master_template_path
from app.schedules import AD_HOC_SCHEDULE_LETTERS, ALL_SCHEDULE_LETTERS
from app.template_config import (
    MasterTemplateMappingV12,
    TemplateMappingFile,
    load_mapping_any,
    schedules_for_matter,
)

_REPO_ROOT = Path(__file__).resolve().parent.parent


def _warn_if_placeholder_master_workbook(wb: Workbook, template_path: Path) -> None:
    """The repo used to ship a tiny stub named like the real doc 2290 file — catch mistaken loads."""
    if "Working Balance" not in wb.sheetnames:
        return
    ws = wb["Working Balance"]
    if (ws.max_row or 0) <= 8 and (ws.max_column or 0) <= 4:
        warnings.warn(
            f"The workbook at {template_path} looks like a minimal stub (Working Balance is "
            f"{ws.max_row}x{ws.max_column}). Use the full firm file "
            f"'2290-Accounting Template.xlsx' and set TEMPLATE_PATH accordingly.",
            UserWarning,
            stacklevel=2,
        )


def _resolve_path(p: Path) -> Path:
    return p if p.is_absolute() else (_REPO_ROOT / p).resolve()


def _col(letter: str) -> int:
    return column_index_from_string(letter.strip().upper())


def _unmerge_covering_cell(ws, row: int, column: int) -> None:
    """Remove a merged range that contains (row, column) so the cell can be written."""
    for rng in list(ws.merged_cells.ranges):
        if (
            rng.min_row <= row <= rng.max_row
            and rng.min_col <= column <= rng.max_col
        ):
            ws.unmerge_cells(str(rng))
            return


def _set_cell_value(ws, row: int, column: int, value: Any) -> None:
    """Write a worksheet cell; unmerge first when the template uses merged placeholders."""
    _unmerge_covering_cell(ws, row, column)
    ws.cell(row=row, column=column, value=value)


def _set_cell_by_a1(ws, cell_ref: str, value: Any) -> None:
    col_letters, row = coordinate_from_string(cell_ref)
    _set_cell_value(ws, row, column_index_from_string(col_letters), value)


def _safe_filename(s: str) -> str:
    out = "".join(c if c.isalnum() or c in "._- " else "_" for c in s)
    return out.strip()[:80] or "Matter"


def parse_schedule_letter(raw: Optional[str]) -> Optional[str]:
    """Return schedule letter for Excel placement, or None if excluded / needs review / not mapped."""
    if not raw:
        return None
    s = str(raw).strip().upper()
    if s in ("INTERNAL_TRANSFER", "EXCLUDED", "NEEDS_REVIEW", "?", ""):
        return None
    if s in ALL_SCHEDULE_LETTERS:
        return s
    return None


def _write_row(
    ws,
    row: int,
    columns: dict[str, str],
    txn: dict[str, Any],
) -> None:
    if "date" in columns and txn.get("txn_date") is not None:
        d = txn["txn_date"]
        val = d.isoformat() if hasattr(d, "isoformat") else str(d)
        ws.cell(row=row, column=_col(columns["date"]), value=val)
    if "description" in columns:
        ws.cell(
            row=row,
            column=_col(columns["description"]),
            value=txn.get("description") or "",
        )
    if "amount" in columns and txn.get("amount") is not None:
        ws.cell(
            row=row,
            column=_col(columns["amount"]),
            value=float(txn["amount"]),
        )
    if "payee" in columns:
        ws.cell(
            row=row,
            column=_col(columns["payee"]),
            value=txn.get("payee") or "",
        )
    if "category" in columns:
        ws.cell(
            row=row,
            column=_col(columns["category"]),
            value=txn.get("subcategory") or "",
        )
    if "notes" in columns:
        ws.cell(
            row=row,
            column=_col(columns["notes"]),
            value=txn.get("notes") or "",
        )


def _generate_legacy(
    matter_type: str,
    matter_name: str,
    period_start: date,
    period_end: date,
    transactions: list[dict[str, Any]],
    statement_by_id: dict[str, dict[str, Any]],
    mapping: TemplateMappingFile,
    verifier_email: Optional[str],
) -> Path:
    template_path = _resolve_path(master_template_path())
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    schedule_map = schedules_for_matter(mapping, matter_type)
    wb = load_workbook(template_path)
    _warn_if_placeholder_master_workbook(wb, template_path)

    by_sched: dict[str, list[dict]] = {k: [] for k in schedule_map}
    for t in transactions:
        if t.get("excluded"):
            continue
        if t.get("internal_transfer"):
            continue
        letter = parse_schedule_letter(t.get("schedule"))
        if not letter or letter not in by_sched:
            continue
        by_sched[letter].append(t)

    written_row: dict[str, int] = {}
    for letter, cfg in schedule_map.items():
        sheet_name = cfg.sheet
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        txns = by_sched.get(letter, [])
        first = cfg.first_data_row
        for i, txn in enumerate(txns):
            r = first + i
            _write_row(ws, r, cfg.columns, txn)
            tid = txn.get("id")
            if tid:
                written_row[tid] = r

    _append_audit_legacy(
        wb,
        transactions,
        statement_by_id,
        written_row,
        verifier_email,
    )
    _append_client_clarification_sheet(wb, transactions, statement_by_id)

    return _save_workbook(wb, matter_name, period_start, period_end)


def _append_audit_legacy(
    wb,
    transactions: list[dict[str, Any]],
    statement_by_id: dict[str, dict[str, Any]],
    written_row: dict[str, int],
    verifier_email: Optional[str],
) -> None:
    audit_sheet = wb.create_sheet("_AuditTrail")
    audit_headers = [
        "Schedule",
        "Subcategory",
        "RowInSchedule",
        "SourceFile",
        "SourcePage",
        "OriginalDescription",
        "NormalizedDescription",
        "Amount",
        "AIScheduleSuggestion",
        "AISubcategorySuggestion",
        "AIConfidence",
        "FinalSchedule",
        "FinalSubcategory",
        "EditedByStaff",
        "VerifiedBy",
        "VerificationTimestamp",
    ]
    for c, h in enumerate(audit_headers, start=1):
        audit_sheet.cell(row=1, column=c, value=h).font = Font(bold=True)

    audit_row = 2
    for t in sorted(
        transactions, key=lambda x: (x.get("statement_id") or "", x.get("id") or "")
    ):
        letter = parse_schedule_letter(t.get("schedule"))
        if t.get("internal_transfer") or t.get("excluded"):
            sched_display = (
                "internal_transfer" if t.get("internal_transfer") else "excluded"
            )
        else:
            sched_display = letter or (t.get("schedule") or "")
        tid = t.get("id")
        rin = written_row.get(tid, "")
        sid = t.get("statement_id")
        st = statement_by_id.get(sid, {})
        amt = t.get("amount")
        amt_val = (
            float(amt)
            if amt is not None and isinstance(amt, (int, float, Decimal))
            else amt
        )
        va = t.get("verified_at")
        row_vals = [
            sched_display,
            t.get("subcategory"),
            rin,
            st.get("original_filename", ""),
            t.get("source_page"),
            t.get("description", ""),
            t.get("normalized_description") or "",
            amt_val,
            t.get("schedule"),
            t.get("subcategory"),
            t.get("confidence"),
            t.get("schedule"),
            t.get("subcategory"),
            bool(t.get("edited_by_staff")),
            verifier_email or t.get("verified_by"),
            va.isoformat() if hasattr(va, "isoformat") else va,
        ]
        for c, val in enumerate(row_vals, start=1):
            audit_sheet.cell(row=audit_row, column=c, value=val)
        audit_row += 1
    audit_sheet.sheet_state = "hidden"


def _append_client_clarification_sheet(
    wb: Workbook,
    transactions: list[dict[str, Any]],
    statement_by_id: dict[str, dict[str, Any]],
) -> None:
    rows = [t for t in transactions if t.get("client_clarification")]
    if not rows:
        return
    ws = wb.create_sheet("Client clarification")
    headers = [
        "Date",
        "Amount",
        "Extracted payee",
        "Payee review",
        "Raw description",
        "AI cleaned",
        "Notes",
        "Statement file",
        "Account last4",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)
    r = 2
    for t in sorted(rows, key=lambda x: (x.get("statement_id") or "", x.get("id") or "")):
        st = statement_by_id.get(t.get("statement_id") or "", {})
        d = t.get("txn_date")
        ws.cell(row=r, column=1, value=_txn_date_for_cell(d) if d else None)
        amt = t.get("amount")
        ws.cell(
            row=r,
            column=2,
            value=float(amt) if amt is not None else None,
        )
        ws.cell(row=r, column=3, value=t.get("payee_raw") or t.get("payee") or "")
        ws.cell(row=r, column=4, value=t.get("payee_normalized") or "")
        ws.cell(row=r, column=5, value=t.get("description") or "")
        ws.cell(row=r, column=6, value=t.get("description_ai_cleaned") or "")
        ws.cell(row=r, column=7, value=t.get("notes") or "")
        ws.cell(row=r, column=8, value=st.get("original_filename") or "")
        ws.cell(row=r, column=9, value=st.get("account_last4") or "")
        r += 1


def _mmddyy(d: date) -> str:
    return f"{d.month}/{d.day}/{str(d.year)[-2:]}"


def _txn_date_for_cell(d: Any) -> Any:
    """Prefer native Excel dates over ISO strings."""
    if d is None:
        return None
    if isinstance(d, datetime):
        return d
    if isinstance(d, date):
        return d
    if isinstance(d, str):
        try:
            return date.fromisoformat(d[:10])
        except ValueError:
            return d
    return d


def _subcategory_data_start_row(meta: Any) -> Optional[int]:
    if not isinstance(meta, dict):
        return None
    if meta.get("dataStartRow") is not None:
        return int(meta["dataStartRow"])
    if meta.get("startRow") is not None:
        return int(meta["startRow"])
    return None


def _resolve_workbook_sheet(wb, name_from_config: Optional[str]) -> Optional[str]:
    """Match sheet tab; prefer exact string (e.g. trailing space in 'Schedule A ')."""
    if not name_from_config:
        return None
    if name_from_config in wb.sheetnames:
        return name_from_config
    n = name_from_config.strip().lower()
    for sn in wb.sheetnames:
        if sn.strip().lower() == n:
            return sn
    return None


def _apply_working_balance_header(
    ws,
    wb_meta: dict[str, Any],
    matter_name: str,
    period_start: date,
    period_end: date,
    session_meta: dict[str, Any],
) -> None:
    """Supports enriched mapping `writeCells` + legacy flat cell keys."""
    write_cells = wb_meta.get("writeCells")
    if isinstance(write_cells, dict):
        ctx = {
            "matterName": matter_name or "",
            "caseNumber": (session_meta.get("case_number") or ""),
            "accountingType": (session_meta.get("accounting_type") or ""),
            "fiduciaryName": (session_meta.get("fiduciary_name") or ""),
            "fiduciaryRole": (session_meta.get("fiduciary_role") or "Conservator"),
            "periodStart": _mmddyy(period_start),
            "periodEnd": _mmddyy(period_end),
        }
        for key, spec in write_cells.items():
            if str(key).startswith("_") or not isinstance(spec, dict):
                continue
            cell_ref = spec.get("cell")
            fmt = spec.get("writeFormat")
            if not cell_ref or fmt is None:
                continue
            try:
                _set_cell_by_a1(ws, cell_ref, str(fmt).format(**ctx))
            except (KeyError, ValueError):
                _set_cell_by_a1(ws, cell_ref, fmt)
        return

    if c := wb_meta.get("matterNameCell"):
        _set_cell_by_a1(ws, c, matter_name)
    if c := wb_meta.get("caseNumberCell"):
        _set_cell_by_a1(ws, c, session_meta.get("case_number") or "")
    if c := wb_meta.get("accountingTypeCell"):
        _set_cell_by_a1(ws, c, session_meta.get("accounting_type") or "")
    if c := wb_meta.get("fiduciaryNameCell"):
        _set_cell_by_a1(ws, c, session_meta.get("fiduciary_name") or "")
    if c := wb_meta.get("periodCell"):
        _set_cell_by_a1(
            ws, c, f"{_mmddyy(period_start)} - {_mmddyy(period_end)}"
        )


def _classify_position_asset_class(p: dict[str, Any]) -> str:
    """Best-effort split: explicit asset_class wins; else infer from description."""
    raw = (p.get("asset_class") or "").strip().lower()
    if raw in ("cash", "non_cash"):
        return raw
    desc = (p.get("security_description") or "").upper()
    sym = (p.get("security_symbol") or "").upper()
    cash_signals = ("MONEY MARKET", "CASH", "SWEEP", "SPAXX", "BANK DEPOSIT")
    if any(sig in desc for sig in cash_signals):
        return "cash"
    if not sym and not (p.get("security_description") or "").strip():
        return "cash"
    return "non_cash"


def _section_data_range_for_positions(section_cfg: dict[str, Any], rows_needed: int) -> tuple[int, int]:
    """Return (data_start_row, data_end_row), expanding the template's default span when needed."""
    start = int(section_cfg.get("dataStartRow", 1))
    end = int(section_cfg.get("dataEndRow", start))
    capacity = end - start + 1
    if rows_needed > capacity:
        end = start + rows_needed - 1
    return start, end


def _write_positions_to_schedule(
    ws,
    schedule_cfg: dict[str, Any],
    positions_for_period: list[dict[str, Any]],
    *,
    valued_as_of: Optional[date],
) -> None:
    """Populate the cash + non_cash sections of Schedule B or Schedule E from positions."""
    if not schedule_cfg:
        return
    cols = schedule_cfg.get("columns") or {}
    if not cols:
        return

    header_cells = schedule_cfg.get("headerCells") or {}
    if valued_as_of and (cell := header_cells.get("valuedAsOfDate")):
        _set_cell_by_a1(ws, cell, valued_as_of.isoformat())

    sections = schedule_cfg.get("sections") or {}
    cash_section = sections.get("cash") or {}
    non_cash_section = sections.get("nonCash") or {}

    grouped: dict[str, list[dict[str, Any]]] = {"cash": [], "non_cash": []}
    for p in positions_for_period:
        grouped[_classify_position_asset_class(p)].append(p)

    def _write_section(section_cfg: dict[str, Any], rows: list[dict[str, Any]]) -> None:
        if not section_cfg or not rows:
            return
        start, end = _section_data_range_for_positions(section_cfg, len(rows))
        for i, p in enumerate(rows):
            r = start + i
            desc_parts = [
                (p.get("security_symbol") or "").strip(),
                (p.get("security_description") or "").strip(),
            ]
            qty = p.get("quantity")
            if qty is not None:
                try:
                    desc_parts.append(f"({float(qty):g} sh)")
                except (TypeError, ValueError):
                    pass
            description = " ".join(part for part in desc_parts if part).strip()
            if "description" in cols:
                _set_cell_value(ws, r, _col(cols["description"]), description)
            mv = p.get("market_value")
            if mv is not None and "marketValue" in cols:
                try:
                    _set_cell_value(ws, r, _col(cols["marketValue"]), float(mv))
                except (TypeError, ValueError):
                    pass
            cb = p.get("cost_basis")
            if cb is None:
                cb = mv  # firm convention: carry value falls back to market value when basis unknown
            if cb is not None and "carryValue" in cols:
                try:
                    _set_cell_value(ws, r, _col(cols["carryValue"]), float(cb))
                except (TypeError, ValueError):
                    pass

        totals_cells = section_cfg.get("totalsCells") or {}
        totals_pattern = section_cfg.get("totalsFormulaPattern") or {}
        for field, cell in totals_cells.items():
            pattern = totals_pattern.get(field)
            if not pattern:
                continue
            try:
                formula = pattern.format(dataStartRow=start, dataEndRow=end)
                _set_cell_by_a1(ws, cell, formula)
            except (KeyError, ValueError):
                continue

    _write_section(cash_section, grouped["cash"])
    _write_section(non_cash_section, grouped["non_cash"])


def _write_schedule_b_e_from_positions(
    wb,
    sched_b_cfg: dict[str, Any],
    sched_e_cfg: dict[str, Any],
    positions: list[dict[str, Any]],
    *,
    statements_order: list[str],
    statement_by_id: dict[str, dict[str, Any]],
) -> None:
    if not positions:
        return
    beginning = [p for p in positions if (p.get("as_of") or "").lower() == "beginning"]
    ending = [p for p in positions if (p.get("as_of") or "").lower() == "ending"]

    period_starts = [
        statement_by_id.get(sid, {}).get("statement_period_start")
        for sid in statements_order
        if statement_by_id.get(sid)
    ]
    period_ends = [
        statement_by_id.get(sid, {}).get("statement_period_end")
        for sid in statements_order
        if statement_by_id.get(sid)
    ]
    earliest_start = min((d for d in period_starts if d), default=None)
    latest_end = max((d for d in period_ends if d), default=None)

    if beginning and sched_b_cfg:
        sheet_name = _resolve_workbook_sheet(wb, sched_b_cfg.get("sheet") or "POH @ Beginning Schedule B")
        if sheet_name:
            _write_positions_to_schedule(
                wb[sheet_name], sched_b_cfg, beginning, valued_as_of=earliest_start
            )
    if ending and sched_e_cfg:
        sheet_name = _resolve_workbook_sheet(wb, sched_e_cfg.get("sheet") or "POH @ End Schedule E")
        if sheet_name:
            _write_positions_to_schedule(
                wb[sheet_name], sched_e_cfg, ending, valued_as_of=latest_end
            )


def _write_brokerage_transactions_sheet(
    wb,
    cfg: dict[str, Any],
    *,
    statements_order: list[str],
    statement_by_id: dict[str, dict[str, Any]],
    transactions: list[dict[str, Any]],
) -> dict[str, tuple[str, int]]:
    """Render brokerage trades on a dedicated 'Brokerage Transactions' sheet."""
    sheet_name = cfg.get("sheet") or "Brokerage Transactions"
    create_if_missing = bool(cfg.get("createIfMissing", True))
    resolved = _resolve_workbook_sheet(wb, sheet_name)
    if not resolved:
        if not create_if_missing:
            return {}
        ws = wb.create_sheet(sheet_name)
        resolved = sheet_name
    else:
        ws = wb[resolved]

    cols = cfg.get("columns") or {
        "date": "B",
        "symbol": "C",
        "description": "D",
        "trade_kind": "E",
        "quantity": "F",
        "price": "G",
        "cost_basis": "H",
        "proceeds": "I",
        "realized_gain_loss": "J",
        "debit": "K",
        "credit": "L",
        "notes": "M",
    }
    headers_cfg = cfg.get("columnHeaders") or {}
    first_block_row = int(cfg.get("firstBlockStartRow", 2))
    block = cfg.get("block") or {}
    stride = int(cfg.get("blockStrideRows") or 16)
    off_name = int(block.get("bankNameRowOffset", 0))
    off_headers = int(block.get("columnHeaderRowOffset", 1))
    off_data = int(block.get("dataStartRowOffset", 2))
    off_data_end = int(block.get("dataEndRowOffset", 8))
    off_totals = int(block.get("totalsRowOffset", 10))
    totals_formulas = cfg.get("totalsFormulas") or {}

    written: dict[str, tuple[str, int]] = {}
    first_data_col = _col(cols.get("date", "B"))

    for block_idx, stmt_id in enumerate(statements_order):
        st = statement_by_id.get(stmt_id)
        if not st:
            continue
        stmt_tx = [t for t in transactions if t.get("statement_id") == stmt_id]
        base = first_block_row + block_idx * stride
        header = " ".join(
            filter(
                None,
                [
                    st.get("institution") or "",
                    st.get("account_last4") and f"…{st.get('account_last4')}",
                ],
            )
        ) or "Brokerage Account"
        _set_cell_value(ws, base + off_name, first_data_col, header)
        header_row = base + off_headers
        for key, col_letter in cols.items():
            label = headers_cfg.get(key) or key.replace("_", " ").title()
            _set_cell_value(ws, header_row, _col(col_letter), label)

        data_start = base + off_data
        data_end = base + off_data_end
        row_ptr = data_start
        for t in stmt_tx:
            if row_ptr > data_end:
                data_end = row_ptr  # expand block when statement has more rows than the template default
            if "date" in cols and t.get("txn_date") is not None:
                _set_cell_value(ws, row_ptr, _col(cols["date"]), _txn_date_for_cell(t["txn_date"]))
            if "symbol" in cols:
                _set_cell_value(ws, row_ptr, _col(cols["symbol"]), t.get("security_symbol") or "")
            if "description" in cols:
                _set_cell_value(ws, row_ptr, _col(cols["description"]), t.get("description") or "")
            if "trade_kind" in cols:
                _set_cell_value(ws, row_ptr, _col(cols["trade_kind"]), t.get("trade_kind") or "")
            for key in ("quantity", "price", "cost_basis", "proceeds", "realized_gain_loss"):
                val = t.get(key)
                if val is None or key not in cols:
                    continue
                try:
                    _set_cell_value(ws, row_ptr, _col(cols[key]), float(val))
                except (TypeError, ValueError):
                    continue
            amt = t.get("amount")
            debit_val = None
            credit_val = None
            if amt is not None:
                try:
                    v = float(amt)
                    if v < 0:
                        debit_val = -v
                    else:
                        credit_val = v
                except (TypeError, ValueError):
                    pass
            if "debit" in cols and debit_val is not None:
                _set_cell_value(ws, row_ptr, _col(cols["debit"]), debit_val)
            if "credit" in cols and credit_val is not None:
                _set_cell_value(ws, row_ptr, _col(cols["credit"]), credit_val)
            if "notes" in cols:
                _set_cell_value(ws, row_ptr, _col(cols["notes"]), t.get("notes") or "")
            tid = t.get("id")
            if tid:
                written[tid] = (resolved, row_ptr)
            row_ptr += 1

        actual_end = max(data_start, row_ptr - 1)
        totals_row = base + off_totals
        if label_cell := totals_formulas.get("totalLabelCell"):
            try:
                _set_cell_by_a1(
                    ws,
                    label_cell.format(totalsRow=totals_row),
                    totals_formulas.get("totalLabelText") or "TOTALS",
                )
            except (KeyError, ValueError):
                pass
        for key, formula in totals_formulas.items():
            if key in ("totalLabelCell", "totalLabelText"):
                continue
            if not isinstance(formula, str) or "{" not in formula:
                continue
            try:
                expanded = formula.format(
                    dataStartRow=data_start,
                    dataEndRow=actual_end,
                    totalsRow=totals_row,
                )
            except (KeyError, ValueError):
                continue
            target_col = None
            if key == "debitTotal":
                target_col = cols.get("debit")
            elif key == "creditTotal":
                target_col = cols.get("credit")
            elif key == "realizedGLTotal":
                target_col = cols.get("realized_gain_loss")
            if target_col:
                _set_cell_value(ws, totals_row, _col(target_col), expanded)

    return written


def _match_subcategory_key(
    label: Optional[str], keys: list[str]
) -> Optional[str]:
    if not label or not keys:
        return None
    L = label.strip().lower()
    for k in keys:
        if k.strip().lower() == L:
            return k
    for k in keys:
        if L in k.lower() or k.lower() in L:
            return k
    return keys[0] if keys else None


def _generate_v12(
    matter_name: str,
    period_start: date,
    period_end: date,
    transactions: list[dict[str, Any]],
    statement_by_id: dict[str, dict[str, Any]],
    statements_order: list[str],
    mapping: MasterTemplateMappingV12,
    verifier_email: Optional[str],
    session_meta: dict[str, Any],
    positions: Optional[list[dict[str, Any]]] = None,
) -> Path:
    settings = get_settings()
    raw_tpl = mapping.template_path or str(settings.template_path)
    template_path = _resolve_path(Path(raw_tpl))
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path)
    _warn_if_placeholder_master_workbook(wb, template_path)

    sheets_cfg = mapping.sheets

    wb_meta = sheets_cfg.get("workingBalance") or {}
    wb_sheet_n = _resolve_workbook_sheet(
        wb, wb_meta.get("sheet") or "Working Balance"
    )
    if wb_sheet_n:
        _apply_working_balance_header(
            wb[wb_sheet_n], wb_meta, matter_name, period_start, period_end, session_meta
        )

    def _is_brokerage(st: dict[str, Any]) -> bool:
        return (st.get("document_type") or "").lower() == "brokerage"

    bank_statements = [
        sid for sid in statements_order
        if (st := statement_by_id.get(sid)) and not _is_brokerage(st)
    ]
    brokerage_statements = [
        sid for sid in statements_order
        if (st := statement_by_id.get(sid)) and _is_brokerage(st)
    ]

    bank_cfg = sheets_cfg.get("bankTransactions") or {}
    bank_sheet_n = _resolve_workbook_sheet(
        wb, bank_cfg.get("sheet") or "Bank Statement Transactions"
    )
    bank_cols = bank_cfg.get("columns") or {
        "date": "A",
        "description": "B",
        "account": "C",
        "check": "D",
        "copy_chk": "E",
        "debit": "F",
        "credit": "G",
        "additional_info": "H",
    }
    first_block_row = int(bank_cfg.get("firstBlockStartRow", 5))
    block = bank_cfg.get("block") or {}
    stride = int(
        bank_cfg.get("blockStrideRows") or bank_cfg.get("blockGapRows") or 16
    )
    off_bank = int(block.get("bankNameRowOffset", 0))
    off_data = int(block.get("dataStartRowOffset", 2))

    written_row: dict[str, tuple[str, int]] = {}  # txn_id -> (sheet name, row)

    if bank_sheet_n:
        ws_bank = wb[bank_sheet_n]
        first_data_col = _col(bank_cols.get("date", "B"))
        for block_idx, stmt_id in enumerate(bank_statements):
            st = statement_by_id.get(stmt_id)
            if not st:
                continue
            stmt_tx = [t for t in transactions if t.get("statement_id") == stmt_id]
            base = first_block_row + block_idx * stride
            header = " ".join(
                filter(
                    None,
                    [
                        st.get("institution") or "",
                        st.get("account_last4")
                        and f"…{st.get('account_last4')}",
                    ],
                )
            )
            name_row = base + off_bank
            _set_cell_value(
                ws_bank,
                name_row,
                first_data_col,
                header or "Account",
            )
            row_ptr = base + off_data
            for t in stmt_tx:
                amt = t.get("amount")
                debit_val = None
                credit_val = None
                if amt is not None:
                    try:
                        v = float(amt)
                        if v < 0:
                            debit_val = -v
                        else:
                            credit_val = v
                    except (TypeError, ValueError):
                        pass
                if "date" in bank_cols and t.get("txn_date") is not None:
                    _set_cell_value(
                        ws_bank,
                        row_ptr,
                        _col(bank_cols["date"]),
                        _txn_date_for_cell(t["txn_date"]),
                    )
                if "description" in bank_cols:
                    _set_cell_value(
                        ws_bank,
                        row_ptr,
                        _col(bank_cols["description"]),
                        t.get("description") or "",
                    )
                if "account" in bank_cols:
                    _set_cell_value(
                        ws_bank,
                        row_ptr,
                        _col(bank_cols["account"]),
                        st.get("account_last4") or "",
                    )
                if "check" in bank_cols:
                    _set_cell_value(
                        ws_bank, row_ptr, _col(bank_cols["check"]), ""
                    )
                if "copy_chk" in bank_cols:
                    _set_cell_value(
                        ws_bank, row_ptr, _col(bank_cols["copy_chk"]), ""
                    )
                if "debit" in bank_cols and debit_val is not None:
                    _set_cell_value(
                        ws_bank,
                        row_ptr,
                        _col(bank_cols["debit"]),
                        debit_val,
                    )
                if "credit" in bank_cols and credit_val is not None:
                    _set_cell_value(
                        ws_bank,
                        row_ptr,
                        _col(bank_cols["credit"]),
                        credit_val,
                    )
                if "additional_info" in bank_cols:
                    _set_cell_value(
                        ws_bank,
                        row_ptr,
                        _col(bank_cols["additional_info"]),
                        t.get("notes") or "",
                    )
                tid = t.get("id")
                if tid:
                    written_row[tid] = (bank_sheet_n, row_ptr)
                row_ptr += 1

    if brokerage_statements:
        broker_written = _write_brokerage_transactions_sheet(
            wb,
            sheets_cfg.get("brokerageTransactions") or {},
            statements_order=brokerage_statements,
            statement_by_id=statement_by_id,
            transactions=transactions,
        )
        written_row.update(broker_written)

    _write_schedule_b_e_from_positions(
        wb,
        sheets_cfg.get("scheduleB") or {},
        sheets_cfg.get("scheduleE") or {},
        positions or [],
        statements_order=statements_order,
        statement_by_id=statement_by_id,
    )

    schedule_a = sheets_cfg.get("scheduleA") or {}
    if schedule_a:
        sheet_a_name = schedule_a.get("sheet") or "Schedule A"
        subs = schedule_a.get("subcategories") or {}
        sub_keys = list(subs.keys())
        cursors: dict[str, int] = {}
        for sk, meta in subs.items():
            sr = _subcategory_data_start_row(meta)
            if sr is not None:
                cursors[sk] = sr

        actual_a = _resolve_workbook_sheet(wb, sheet_a_name)
        if actual_a:
            ws_a = wb[actual_a]
            for t in transactions:
                if t.get("excluded") or t.get("internal_transfer"):
                    continue
                if parse_schedule_letter(t.get("schedule")) != "A":
                    continue
                sk = _match_subcategory_key(t.get("subcategory"), sub_keys)
                if not sk or sk not in cursors:
                    continue
                r = cursors[sk]
                cols = schedule_a.get("columns") or {
                    "date": "A",
                    "description": "B",
                    "amount": "C",
                }
                if "date" in cols and t.get("txn_date") is not None:
                    _set_cell_value(
                        ws_a,
                        r,
                        _col(cols["date"]),
                        _txn_date_for_cell(t["txn_date"]),
                    )
                if "description" in cols:
                    _set_cell_value(
                        ws_a,
                        r,
                        _col(cols["description"]),
                        t.get("description") or "",
                    )
                if "amount" in cols and t.get("amount") is not None:
                    _set_cell_value(
                        ws_a, r, _col(cols["amount"]), float(t["amount"])
                    )
                tid = t.get("id")
                if tid:
                    written_row[tid] = (actual_a, r)
                cursors[sk] = r + 1

    schedule_c = sheets_cfg.get("scheduleC") or {}
    if schedule_c:
        sheet_c_name = schedule_c.get("sheet") or "Schedule C"
        subs_c = schedule_c.get("subcategories") or {}
        sub_keys_c = list(subs_c.keys())
        cursors_c: dict[str, int] = {}
        for sk, meta in subs_c.items():
            sr = _subcategory_data_start_row(meta)
            if sr is not None:
                cursors_c[sk] = sr

        actual_c = _resolve_workbook_sheet(wb, sheet_c_name)
        if actual_c:
            ws_c = wb[actual_c]
            cols_c = schedule_c.get("columns") or {
                "date": "A",
                "account": "B",
                "description": "C",
                "check": "D",
                "amount": "E",
            }
            for t in transactions:
                if t.get("excluded") or t.get("internal_transfer"):
                    continue
                if parse_schedule_letter(t.get("schedule")) != "C":
                    continue
                sk = _match_subcategory_key(t.get("subcategory"), sub_keys_c)
                if not sk or sk not in cursors_c:
                    continue
                r = cursors_c[sk]
                st = statement_by_id.get(t.get("statement_id") or "", {})
                if "date" in cols_c and t.get("txn_date") is not None:
                    _set_cell_value(
                        ws_c,
                        r,
                        _col(cols_c["date"]),
                        _txn_date_for_cell(t["txn_date"]),
                    )
                if "account" in cols_c:
                    _set_cell_value(
                        ws_c,
                        r,
                        _col(cols_c["account"]),
                        st.get("account_last4") or "",
                    )
                if "description" in cols_c:
                    _set_cell_value(
                        ws_c,
                        r,
                        _col(cols_c["description"]),
                        t.get("description") or "",
                    )
                if "check" in cols_c:
                    _set_cell_value(ws_c, r, _col(cols_c["check"]), "")
                if "amount" in cols_c and t.get("amount") is not None:
                    _set_cell_value(
                        ws_c, r, _col(cols_c["amount"]), float(t["amount"])
                    )
                tid = t.get("id")
                if tid:
                    written_row[tid] = (actual_c, r)
                cursors_c[sk] = r + 1

    schedule_f = sheets_cfg.get("scheduleF") or {}
    if schedule_f:
        sheet_f_name = schedule_f.get("sheet") or "Schedule F"
        data_start = int(schedule_f.get("dataStartRow", 7))
        cols_f = schedule_f.get("columns") or {
            "date": "A",
            "description": "B",
            "carry_value": "C",
        }

        actual_f = _resolve_workbook_sheet(wb, sheet_f_name)
        if actual_f:
            ws_f = wb[actual_f]
            r = data_start
            for t in transactions:
                if t.get("excluded") or t.get("internal_transfer"):
                    continue
                if parse_schedule_letter(t.get("schedule")) != "F":
                    continue
                if "date" in cols_f and t.get("txn_date") is not None:
                    _set_cell_value(
                        ws_f,
                        r,
                        _col(cols_f["date"]),
                        _txn_date_for_cell(t["txn_date"]),
                    )
                if "description" in cols_f:
                    _set_cell_value(
                        ws_f,
                        r,
                        _col(cols_f["description"]),
                        t.get("description") or "",
                    )
                key_cv = "carry_value" if "carry_value" in cols_f else "amount"
                if key_cv in cols_f and t.get("amount") is not None:
                    _set_cell_value(
                        ws_f,
                        r,
                        _col(cols_f[key_cv]),
                        float(t["amount"]),
                    )
                tid = t.get("id")
                if tid:
                    written_row[tid] = (actual_f, r)
                r += 1

    for letter in AD_HOC_SCHEDULE_LETTERS:
        ad_tx = []
        for t in transactions:
            if t.get("excluded") or t.get("internal_transfer"):
                continue
            if parse_schedule_letter(t.get("schedule")) == letter:
                ad_tx.append(t)
        if not ad_tx:
            continue
        meta = mapping.ad_hoc_schedules.get(letter)
        label = meta.label if meta else f"Schedule {letter}"
        sheet_title = f"Schedule {letter}"
        if sheet_title in wb.sheetnames:
            ws_ad = wb[sheet_title]
        else:
            ws_ad = wb.create_sheet(sheet_title)
            ws_ad["A1"] = label
            ws_ad["A2"] = "Date"
            ws_ad["B2"] = "Description"
            ws_ad["C2"] = "Amount"
            ws_ad["D2"] = "Subcategory"
        r = 3
        for t in ad_tx:
            if t.get("txn_date") is not None:
                _set_cell_value(
                    ws_ad,
                    r,
                    1,
                    _txn_date_for_cell(t["txn_date"]),
                )
            _set_cell_value(ws_ad, r, 2, t.get("description") or "")
            if t.get("amount") is not None:
                _set_cell_value(ws_ad, r, 3, float(t["amount"]))
            _set_cell_value(ws_ad, r, 4, t.get("subcategory") or "")
            tid = t.get("id")
            if tid:
                written_row[tid] = (sheet_title, r)
            r += 1

    _append_audit_v12(
        wb,
        transactions,
        statement_by_id,
        written_row,
        verifier_email,
    )
    _append_client_clarification_sheet(wb, transactions, statement_by_id)

    return _save_workbook(wb, matter_name, period_start, period_end)


def _decimal_to_float(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, Decimal):
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return v


def _append_audit_v12(
    wb,
    transactions: list[dict[str, Any]],
    statement_by_id: dict[str, dict[str, Any]],
    written_row: dict[str, tuple[str, int]],
    verifier_email: Optional[str],
) -> None:
    audit_sheet = wb.create_sheet("_AuditTrail")
    headers = [
        "Schedule",
        "Subcategory",
        "RowInScheduleSheet",
        "SourceFile",
        "SourcePage",
        "OriginalDescription",
        "NormalizedDescription",
        "Amount",
        "TradeKind",
        "Symbol",
        "Quantity",
        "Price",
        "CostBasis",
        "Proceeds",
        "RealizedGainLoss",
        "AIScheduleSuggestion",
        "AISubcategorySuggestion",
        "AIConfidence",
        "FinalSchedule",
        "FinalSubcategory",
        "EditedByStaff",
        "VerifiedBy",
        "VerificationTimestamp",
    ]
    for c, h in enumerate(headers, start=1):
        audit_sheet.cell(row=1, column=c, value=h).font = Font(bold=True)

    row_i = 2
    for t in sorted(
        transactions, key=lambda x: (x.get("statement_id") or "", x.get("id") or "")
    ):
        letter = parse_schedule_letter(t.get("schedule"))
        if t.get("internal_transfer") or t.get("excluded"):
            sched_display = (
                "internal_transfer" if t.get("internal_transfer") else "excluded"
            )
        else:
            sched_display = letter or (t.get("schedule") or "")
        tid = t.get("id")
        loc = written_row.get(tid)
        rin = loc[1] if loc else ""
        sid = t.get("statement_id")
        st = statement_by_id.get(sid, {})
        va = t.get("verified_at")
        vals = [
            sched_display,
            t.get("subcategory"),
            rin,
            st.get("original_filename", ""),
            t.get("source_page"),
            t.get("description", ""),
            t.get("normalized_description") or "",
            _decimal_to_float(t.get("amount")),
            t.get("trade_kind") or "",
            t.get("security_symbol") or "",
            _decimal_to_float(t.get("quantity")),
            _decimal_to_float(t.get("price")),
            _decimal_to_float(t.get("cost_basis")),
            _decimal_to_float(t.get("proceeds")),
            _decimal_to_float(t.get("realized_gain_loss")),
            t.get("schedule"),
            t.get("subcategory"),
            t.get("confidence"),
            t.get("schedule"),
            t.get("subcategory"),
            bool(t.get("edited_by_staff")),
            verifier_email or t.get("verified_by"),
            va.isoformat() if hasattr(va, "isoformat") else va,
        ]
        for c, val in enumerate(vals, start=1):
            audit_sheet.cell(row=row_i, column=c, value=val)
        row_i += 1
    audit_sheet.sheet_state = "hidden"


def _save_workbook(
    wb: Workbook, matter_name: str, period_start: date, period_end: date
) -> Path:
    out_name = f"{_safe_filename(matter_name)}_Accounting_{period_start}_to_{period_end}_{date.today()}.xlsx"
    out_path = _REPO_ROOT / "data" / "exports" / out_name
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def generate_accounting_workbook(
    matter_type: str,
    matter_name: str,
    period_start: date,
    period_end: date,
    transactions: list[dict[str, Any]],
    statement_by_id: dict[str, dict[str, Any]],
    mapping_path: Path,
    verifier_email: Optional[str],
    statement_order: Optional[list[str]] = None,
    session_meta: Optional[dict[str, Any]] = None,
    positions: Optional[list[dict[str, Any]]] = None,
) -> Path:
    """
    Generate Excel export. Uses spec v1.2 mapping when JSON contains `sheets.workingBalance`;
    otherwise legacy flat schedule→sheet mapping.

    positions (optional): holdings rows (period start + end) used to populate Schedule B /
    Schedule E on brokerage statements. Each row has keys: statement_id, as_of, asset_class,
    security_symbol, security_description, quantity, unit_price, market_value, cost_basis.
    """
    meta = session_meta or {}
    mapping = load_mapping_any(mapping_path)
    stmt_order = statement_order or list(statement_by_id.keys())
    positions_list = positions or []

    if isinstance(mapping, MasterTemplateMappingV12):
        return _generate_v12(
            matter_name=matter_name,
            period_start=period_start,
            period_end=period_end,
            transactions=transactions,
            statement_by_id=statement_by_id,
            statements_order=stmt_order,
            mapping=mapping,
            verifier_email=verifier_email,
            session_meta=meta,
            positions=positions_list,
        )

    return _generate_legacy(
        matter_type=matter_type,
        matter_name=matter_name,
        period_start=period_start,
        period_end=period_end,
        transactions=transactions,
        statement_by_id=statement_by_id,
        mapping=mapping,
        verifier_email=verifier_email,
    )
